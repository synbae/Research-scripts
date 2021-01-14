import openpyxl

path = './Files/movies.xlsx'

workbook = openpyxl.load_workbook(path)
sheet = workbook.active             #workbook.get_sheet_by_name("sheet1")

rows = sheet.max_row
cols = sheet.max_column

print(rows, cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r, column=c).value, end=" ")
    print()